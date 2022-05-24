import gspread
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime
import sys
import os
import win32com.client
import apiclient.discovery
import httplib2
import oauth2client.file
import oauth2client.tools
import re
import requests
import shutil
import urllib.parse

mydate = datetime.datetime.now()
month = mydate.strftime("%b") 


#Converts google sheet to csv file and saves to downloads folder.
SCOPES = 'https://www.googleapis.com/auth/drive.readonly'
SPREADSHEET_ID = '1YVynNUsRZ7y-jXZZbHHMdXZkKqgiXHVgsQiQRZqwdHM'

store = oauth2client.file.Storage('credentials.json')
creds = store.get()
if not creds or creds.invalid:
  flow = oauth2client.client.flow_from_clientsecrets('C:\\Users\\ckallenberger\\pycharmprojects\\Travelsheetautomation\\client_secret.json', SCOPES)
  creds = oauth2client.tools.run_flow(flow, store)

service = apiclient.discovery.build('sheets', 'v4', http=creds.authorize(httplib2.Http()))

result = service.spreadsheets().get(spreadsheetId = SPREADSHEET_ID).execute()
urlParts = urllib.parse.urlparse(result['spreadsheetUrl'])
path = re.sub("\/edit$", '/export', urlParts.path)
urlParts = urlParts._replace(path=path)
headers = {
  'Authorization': 'Bearer ' + creds.access_token,
}
for sheet in result['sheets']:
  params = {
    'id': SPREADSHEET_ID,
    'format': 'xlsx',
    'gid': sheet['properties']['sheetId'],
  }
  queryParams = urllib.parse.urlencode(params)
  urlParts = urlParts._replace(query=queryParams)
  url = urllib.parse.urlunparse(urlParts)
  response = requests.get(url, headers = headers)
  filePath = f'C:\\Users\\ckallenberger\\downloads\\TravelForms\\TravelSheet{month}.xlsx' 
  with open(filePath, 'wb') as csvFile:
    csvFile.write(response.content)


#Grabs value of mileage cell and saves it in the mileage chart excel file.
gc = gspread.service_account()
sht1 = gc.open_by_key('1YVynNUsRZ7y-jXZZbHHMdXZkKqgiXHVgsQiQRZqwdHM')
wks = sht1.get_worksheet(0)
mileage = wks.acell('M35').value
workbook = load_workbook(filename=r"C:\Users\ckallenberger\Downloads\Travel Form Jan 2022.xlsx")
sheet = workbook.active
sheet["L11"] = mileage
workbook.save(f"C:\\Users\\ckallenberger\\Downloads\\TravelForms\\Travel Form {month} 2022.xlsx")

workbook = load_workbook(filename=f'C:\\Users\\ckallenberger\\downloads\\TravelForms\\TravelSheet{month}.xlsx')
ws=workbook.active
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
for col, value in dims.items():
    ws.column_dimensions[col].width = value



o = win32com.client.Dispatch("Excel.Application")

o.Visible = False

wb_path = f'c:\\users\\ckallenberger\\downloads\\TravelForms\\TravelSheet{month}.xlsx'

wb = o.Workbooks.Open(wb_path)



ws_index_list = [1] #say you want to print these sheets

path_to_pdf = r'C:\users\ckallenberger\downloads\TravelForms\sample.pdf'



wb.Worksheets(ws_index_list).Select()

wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)

print_area = 'A1:M35'



for index in ws_index_list:

    #off-by-one so the user can start numbering the worksheets at 1

    ws = wb.Worksheets[index]

    ws.PageSetup.Zoom = False

    ws.PageSetup.FitToPagesTall = 1

    ws.PageSetup.FitToPagesWide = 1

    ws.PageSetup.PrintArea = print_area


wb.Close(True)


